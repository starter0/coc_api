#! /usr/bin/env python3


#######################################################
### Get Clan war league information              ######
### Convert info to excel file 			####### 
#######################################################


import os, sys, json
import subprocess
import datetime, time
import requests

import openpyxl
from openpyxl.styles import Font, Alignment

#######################################################
############### Define Variable #######################
#######################################################

MY_CLAN_TAG='#982JCPJU'
API_BASE_URL = 'https://api.clashofclans.com/v1'
MY_KEY2= "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6ImYzNjM2NDFiLWRkMTEtNDkwOS1hOGRmLTlhMzAyMmUyYmE2YSIsImlhdCI6MTU4NDQwNDg3OSwic3ViIjoiZGV2ZWxvcGVyLzY5ZDVjYmRiLTg5NTQtMDc4Zi1hNzYwLTY2ZmY4MDFiMTc3MyIsInNjb3BlcyI6WyJjbGFzaCJdLCJsaW1pdHMiOlt7InRpZXIiOiJkZXZlbG9wZXIvc2lsdmVyIiwidHlwZSI6InRocm90dGxpbmcifSx7ImNpZHJzIjpbIjE4Mi4yMjkuODAuMjAiXSwidHlwZSI6ImNsaWVudCJ9XX0.ehu1yYEbd7NkbuEJ6NGHF1AzRXcY9tTTEvbkr8xKdMlzTMqJ_tkq11LnBUsQwAefMaG-cPz0DvWJMRI3p0j7Rg"

def member_arrange(member_obj) :

	arranged_member_list = {}

	for member in member_obj['members'] :
		arranged_member_list.update({member['tag'] : member['mapPosition']} )

	return (arranged_member_list)		

def xlsx_idle_clan_war_league(path, filename_xls, op_clan, round) :
	
    if os.path.isfile(path+ filename_xls) :
        wb = openpyxl.load_workbook(path+filename_xls)
    else :
        wb = openpyxl.Workbook()


    sh_name = str(datetime.date.today().month)+'_round_'+str(round)
    
    sh = wb.create_sheet(sh_name, 0)
    sh['B1'] = 'VS' 
    sh['C1'] = op_clan
    sh.merge_cells('C1:D1')
    sh['B2'] = 'Member Name' 
    sh['C2'] = 'Position'
    sh['D2'] = '1th attack position'
    sh['E2'] = 'stars'
    sh['F2'] = 'Destruction Percentage'
    sh['G2'] = 'Not used attack count'
    sh['H2'] = 'Total stars'
    sh['I2'] = 'Town hall level'
    sh.merge_cells('M1:O1')
    sh['J1'] = 'Best opponent attack'
    sh['J2'] = 'Attacker position'
    sh['K2'] = 'stars'
    sh['L2'] = 'DestructionPercentage' 
    sh['M2'] = 'opponentAttacks'

    # style
    sh.freeze_panes = 'A4'
    sh['C1'].font = Font(size = 20, bold = True)
    sh['B2'].font = Font(bold = True)
    sh.column_dimensions['A'].width = 22
    sh.column_dimensions['B'].width = 17
    sh.column_dimensions['D'].width = 17
    sh.column_dimensions['F'].width = 22
    sh.column_dimensions['G'].width = 22
    sh.column_dimensions['I'].width = 15
    sh.column_dimensions['J'].width = 22
    sh.column_dimensions['L'].width = 22
    sh.column_dimensions['M'].width = 17

    wb.save(path+filename_xls)
"""
    rows = sh.max_row
    cols = sh.max_column
    for r in range(1, rows) :
        for c in range(0,cols) :
            sh.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sh.cell(row = r+1, column = c+1).font = Font(bold = True)

"""
	

def append_xlsx(path, obj, filename_xls) :
    wb = openpyxl.load_workbook(path + filename_xls)
    sh = wb.active
    sh.append(obj)
    wb.save(path + filename_xls)


def kj_api_login() :
	coc_id = '80dack@naver.com'
	coc_passwd = 'dgenius11'

	MY_KEY2= "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6ImYzNjM2NDFiLWRkMTEtNDkwOS1hOGRmLTlhMzAyMmUyYmE2YSIsImlhdCI6MTU4NDQwNDg3OSwic3ViIjoiZGV2ZWxvcGVyLzY5ZDVjYmRiLTg5NTQtMDc4Zi1hNzYwLTY2ZmY4MDFiMTc3MyIsInNjb3BlcyI6WyJjbGFzaCJdLCJsaW1pdHMiOlt7InRpZXIiOiJkZXZlbG9wZXIvc2lsdmVyIiwidHlwZSI6InRocm90dGxpbmcifSx7ImNpZHJzIjpbIjE4Mi4yMjkuODAuMjAiXSwidHlwZSI6ImNsaWVudCJ9XX0.ehu1yYEbd7NkbuEJ6NGHF1AzRXcY9tTTEvbkr8xKdMlzTMqJ_tkq11LnBUsQwAefMaG-cPz0DvWJMRI3p0j7Rg"

	API_PAGE_BASE = "https://developer.clashofclans.com/"
	login_page = "https://developer.clashofclans.com/api/login"

	login_data = {"email": coc_id, "password": coc_passwd}
	headers = {"content-type": "application/json"}

	response = requests.post(login_page, headers = headers, json = login_data)

	return response



def kj_log_set() :
	mylogger = logging.getLogger("kj_coc") #create kj_coc logging
	mylogger.setLevel(logging.DEBUG) # more than DEBUG

	formatter = logging.Formatter('%(asctime)s > %(module)s\t%(levelname)s\t%(message)s', '%Y-%m-%d %H:%M:%S')

	file_handler = logging.FileHandler('./kj_coc_war_league_log.txt')
	file_handler.setFormatter(formatter)
	file_handler.setLevel(logging.DEBUG) # more than DEBUG
	mylogger.addHandler(file_handler)

	mylogger = logging.getLogger("kj_coc") #create kj_coc logging
	mylogger.setLevel(logging.DEBUG) # more than DEBUG

	formatter = logging.Formatter('%(asctime)s > %(module)s\t%(levelname)s\t%(message)s', '%Y-%m-%d %H:%M:%S')

	file_handler = logging.FileHandler('./kj_coc_log.txt')
	file_handler.setFormatter(formatter)
	file_handler.setLevel(logging.DEBUG) # more than DEBUG
	mylogger.addHandler(file_handler)


def get_clanwar_league () :

	auth = { 'Authorization': 'Bearer {}'.format(MY_KEY2) }
	request_url = ('https://api.clashofclans.com/v1/clans/{clan_tag}/currentwar'.format(clan_tag=requests.utils.quote(my_clan_tag)) )
	s = requests.Session()

	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %MY_KEY2})

	return res


def get_war_result (wartag) :

	auth = { 'Authorization': 'Bearer {}'.format(MY_KEY2) }
	request_url = ('https://api.clashofclans.com/v1/clanwarleagues/wars/{warTag}'.format(warTag=requests.utils.quote(wartag)) )
	s = requests.Session()

	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %MY_KEY2})

	return res
	

def get_currentwar_league (clantag) :
	
	auth = { 'Authorization': 'Bearer {}'.format(MY_KEY2) }
	get_cu_war_league_url = '/clans/{clanTag}/currentwar/leaguegroup'.format(clanTag = requests.utils.quote(clantag))
	
	request_url = API_BASE_URL + get_cu_war_league_url

	s = requests.Session()
	
	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %MY_KEY2})

	return res

def get_league_result ( my_clan, op_clan, round) :


	opponent_members = member_arrange(op_clan)
	my_members = member_arrange(my_clan)

	opponent_clan_name = op_clan['name']


	position_arrange(opponent_members)
	position_arrange(my_members)
	
	title = 'Clan_war_league_'+str(datetime.date.today())+'.xlsx'
	

	xlsx_idle_clan_war_league('./',title,opponent_clan_name, round)
	

	## war result

	for member in my_clan['members'] :

		append_obj = ['']

		append_obj.append(member['name'])
		append_obj.append(my_members[member['tag']])

		
		if ( 'attacks' in member ) :
			for attack in member['attacks'] :
				#find opponent postion
				append_obj.append(opponent_members[(attack['defenderTag'])])
				append_obj.append(attack['stars'])
				append_obj.append(attack['destructionPercentage'])

	
		# caculate miised attack count
		if ( len (append_obj) < 4 ) : # If not attacked
			remain_count = 1
			total_stars = 0
			append_obj.extend(('','',''))

		else :
			remain_count = 0
			total_stars = int(append_obj[4]) 

		append_obj.append(remain_count)

		# Total stars

		append_obj.append(total_stars)
		
		# town hall
		append_obj.append(member['townhallLevel'])
		
		### Best opponent attack info 
		#if ( 'opponentAttacks' in member ) :
		if ( int(member['opponentAttacks']) >= 1 ) :
			#opponent_position = opponent_members[member['bestOpponentAttack']['attackerTag']]
			#append_obj.append(opponent_member[member['bestOpponentAttack']['attackerTag']])
			append_obj.append(opponent_members[member['bestOpponentAttack']['attackerTag']])
			append_obj.append(member['bestOpponentAttack']['stars'])
			append_obj.append(member['bestOpponentAttack']['destructionPercentage'])
			append_obj.append(member['opponentAttacks'])

		else : 
			append_obj.extend(('','','',''))

		append_xlsx('./',append_obj,title)
		#print (append_obj)
	
	return title

def position_arrange(members) :
	
	new_list = sorted(members.items(), key= lambda item: item[1])
	
	for i, key in enumerate(new_list) :
		for member in members.keys() :
			if ( key[0] == member ) :
				members[member] = i + 1
	
def set_cell_value(filename_xls, location, value) :
	wb = openpyxl.load_workbook(filename_xls)
	sh = wb.active
	sh[location] = value 
	wb.save(filename_xls)

	
	
 
if __name__ == '__main__' :

	login_response = kj_api_login()

	print ('Try to get current war league ')
	cu_league_response = get_currentwar_league (MY_CLAN_TAG)

	if (cu_league_response.status_code != 200) : 
		print ('Error Code is ',cu_league_response.status_code)
		exit (1)

	cu_league_txt = json.loads(cu_league_response.content.decode('utf-8'))	

#	with open ('WAR_LEAGE_0310.json') as json_file :
#		json_data = json.load(json_file)
#		cu_league_txt = json_data

	with open (str(datetime.date.today().month)+'_clanwar_league.json') as json_file: 
		json.dump(cu_league_txt, json_file, indent='\t') 


	#### extract wartag
	for i, round in enumerate(cu_league_txt['rounds']) :
		
		for wartag in round['warTags'] :
		
			cu_war_result = get_war_result(wartag)
			
			cu_war_result_txt = json.loads(cu_war_result.content.decode('utf-8'))
 
			if (cu_war_result_txt['state'] == 'warEnded' ) :

				if  (cu_war_result_txt['clan']['tag'] == MY_CLAN_TAG):
					print ('Try to get league result of round ',i+1 )
					result_excel_filename = get_league_result(cu_war_result_txt['clan'], cu_war_result_txt['opponent'], i+1)
					set_cell_value(result_excel_filename, 'A7', cu_war_result_txt['opponent']['clanLevel'])
					set_cell_value(result_excel_filename, 'A10', cu_war_result_txt['clan']['stars'])
					set_cell_value(result_excel_filename, 'A13', cu_war_result_txt['clan']['destructionPercentage'])
					


				elif (cu_war_result_txt['opponent']['tag'] == MY_CLAN_TAG):
					print ('Try to get league result of round ',i+1 )
					result_excel_filename = get_league_result(cu_war_result_txt['opponent'], cu_war_result_txt['clan'], i+1)
					set_cell_value(result_excel_filename, 'A7', cu_war_result_txt['clan']['clanLevel'])
					set_cell_value(result_excel_filename, 'A10', cu_war_result_txt['opponent']['stars'])
					set_cell_value(result_excel_filename, 'A13', cu_war_result_txt['opponent']['destructionPercentage'])
					

				else :
					continue

				end_time = cu_war_result_txt['endTime'][:8]
				set_cell_value(result_excel_filename, 'A4', '20{year}-{month}-{day}'.format(year=end_time[:4], month =end_time[4:6], day = end_time[6:8] ) )
				set_cell_value(result_excel_filename, 'A3', 'Date')
				set_cell_value(result_excel_filename, 'A6', 'Opponent Clan Level')
				set_cell_value(result_excel_filename, 'A9', 'Stars')
				set_cell_value(result_excel_filename, 'A12', 'Destruction Percentage')
					
				


			
	
