#! /usr/bin/env python3

import os, sys, json
import subprocess
import datetime, time

import openpyxl
from openpyxl.styles import Font, Alignment


def member_arrange(member_obj) :

	arranged_member_list = {}

	for member in member_obj['members'] :
		arranged_member_list.update({member['tag'] : member['mapPosition']} )

	return (arranged_member_list)		

def xlsx_idle(path, filename_xls, op_clan) :
	
    if os.path.isfile(path+ filename_xls) :
    #if os.path.isfile(path+'mem_info.xlsx') :
        wb = openpyxl.load_workbook(path+filename_xls)
        #wb = openpyxl.load_workbook(path+'mem_info.xlsx')
    else :
        wb = openpyxl.Workbook()
    
    sh = wb.create_sheet(str(datetime.datetime.now().date()), 0)
    sh['B1'] = 'VS' 
    sh['C1'] = op_clan
    sh.merge_cells('C1:D1')
    sh['B2'] = 'Member Name' 
    sh['C2'] = 'Position'
    sh['D2'] = '1th attack position'
    sh['E2'] = 'stars'
    sh['F2'] = 'Destruction Percentage'
    sh['G2'] = '2st attack position'
    sh['H2'] = 'stars'
    sh['I2'] = 'DestructionPercentage'
    sh['J2'] = 'Not used attack count'
    sh['K2'] = 'Total stars'
    sh['L2'] = 'Town hall level'
    sh.merge_cells('M1:O1')
    sh['M1'] = 'Best opponent attack'
    sh['M2'] = 'Attacker position'
    sh['N2'] = 'stars'
    sh['O2'] = 'DestructionPercentage' 
    sh['P2'] = 'opponentAttacks'
    #sh['L3'] = 'Baba king'
    #sh['M3'] = 'Archer queen'
    #sh['N3'] = 'Warden'
    #sh['O3'] = 'Royal champion'
    
    
    # style
    #sh.freeze_panes = 'A4'
    sh['B1'].font = Font(size = 20, bold = True)

	
    #set cell width
    sh.column_dimensions['A'].width = 22
    sh.column_dimensions['B'].width = 17


    wb.save(path+filename_xls)
"""
    rows = sh.max_row
    cols = sh.max_column
    for r in range(1, rows) :
        for c in range(0,cols) :
            sh.cell(row = r+1, column = c+1).alignment = Alignment(horizontal='center', vertical='center')
            sh.cell(row = r+1, column = c+1).font = Font(bold = True)

"""
	

def append_xlsx(path, obj, filename_xls) :
    wb = openpyxl.load_workbook(path + filename_xls)
    sh = wb.active
    sh.append(obj)
    wb.save(path + filename_xls)


def war_to_excel(war_result) :

	#set member and postion number
	opponent_members = member_arrange(war_result['opponent'])
	opponent_clan_name = war_result['opponent']['name']
	my_clan_members = member_arrange(war_result['clan']) 

	# sort member postion
	position_arrange(opponent_members)
	position_arrange(my_clan_members)

		
	# set excel file name
	title = 'Clan_war_'+str(datetime.date.today().month)+'.xlsx'	
	#title = 'Clan_war_'+str(datetime.date.today().month)+'_month.xlsx'	

	xlsx_idle('./',title,opponent_clan_name)

	for member in war_result['clan']['members'] :
		append_obj = ['']
		# set member's name, postion
		append_obj.append(member['name'])
		append_obj.append(my_clan_members[member['tag']])

		
		if ( 'attacks' in member ) :
			for attack in member['attacks'] :
				#find opponent postion
				append_obj.append(opponent_members[(attack['defenderTag'])])
				append_obj.append(attack['stars'])
				append_obj.append(attack['destructionPercentage'])

	
		# caculate miised attack count
		if ( len (append_obj) < 4 ) : # If not attacked
			remain_count = 2
			total_stars = 0
			append_obj.extend(('','','','','',''))

		elif ( len (append_obj) < 7 ) : # If attacked 1 time
			remain_count = 1
			total_stars = append_obj[4]
			append_obj.extend(('','',''))
		else :
			remain_count = 0
			total_stars = int(append_obj[4]) + int(append_obj[7])

		append_obj.append(remain_count)

		# Total stars

		append_obj.append(total_stars)
		
		# town hall
		append_obj.append(member['townhallLevel'])
		
		### Best opponent attack info 
		if ( 'opponentAttacks' in member ) :
			#opponent_position = opponent_members[member['bestOpponentAttack']['attackerTag']]
			append_obj.append(opponent_members[member['bestOpponentAttack']['attackerTag']])
			append_obj.append(member['bestOpponentAttack']['stars'])
			append_obj.append(member['bestOpponentAttack']['destructionPercentage'])
			append_obj.append(member['opponentAttacks'])

		else : 
			append_obj.extend('','','','')

		append_xlsx('./',append_obj,title)

		append_obj = ['']

	## summary information set
	end_time = war_result['endTime'][:8]
	set_cell_value(title, 'A3', 'End Date')	
	set_cell_value(title, 'A4','{year}-{month}-{day}'.format(year=end_time[:4], month =end_time[4:6], day = end_time[6:8] ) )

	set_cell_value(title, 'A6', 'Opponent Clan Level')	
	set_cell_value(title, 'A7', war_result['opponent']['clanLevel'])	

	set_cell_value(title, 'A9', 'Stars')	
	set_cell_value(title, 'A10', war_result['clan']['stars'])	
	
	set_cell_value(title, 'A12', 'Destruction Percentage')	
	set_cell_value(title, 'A13', war_result['clan']['destructionPercentage'])	

	
	set_cell_value(title, 'A15', 'Result')
	my_clan_stars =  int(war_result['clan']['stars'])
	opp_clan_stars =  int(war_result['opponent']['stars'])
	my_clan_destruction = int(war_result['clan']['destructionPercentage'])
	opp_clan_destruction = int(war_result['opponent']['destructionPercentage'])


	if ( my_clan_stars > opp_clan_stars) :
		result_txt = 'Win'
	elif (my_clan_stars < opp_clan_stars) :
		result_txt = 'Lose'
	elif (my_clan_destruction > opp_clan_destruction) :
		result_txt = 'Win'
	else :
		result_txt = 'Lose'
	set_cell_value(title, 'A16', result_txt)	


def position_arrange(members) :

	new_list = sorted(members.items(), key= lambda item: item[1])

	for i, key in enumerate(new_list) :
		for member in members.keys() :
			if ( key[0] == member ) :
				members[member] = i + 1
				break


def set_cell_value(filename_xls, location, value) :
	wb = openpyxl.load_workbook(filename_xls)
	sh = wb.active
	sh[location] = value
	wb.save(filename_xls)


if __name__ == '__main__' :

	if ( len(sys.argv) < 2) :
		print ('Usage is')
		exit (1)	

	with open(sys.argv[1]) as json_file:
		json_data = json.load(json_file)

	war_to_excel(json_data)	

"""
	opponent_member = member_arrange(json_data['opponent'])
	opponent_clan_name = json_data['opponent']['name']

	title = 'Clan_war.xlsx'	

	xlsx_idle('./',title,opponent_clan_name)
	append_obj = ['']

	for member in json_data['clan']['members'] :
		append_obj.append(member['name'])
		append_obj.append(member['mapPosition'])

		
		if ( 'attacks' in member ) :
			for attack in member['attacks'] :
				#find opponent postion
				append_obj.append(opponent_member[(attack['defenderTag'])])
				append_obj.append(attack['stars'])
				append_obj.append(attack['destructionPercentage'])

	
		# caculate miised attack count
		if ( len (append_obj) < 4 ) : # If not attacked
			remain_count = 2
			total_stars = 0
			append_obj.extend(('','','','','',''))

		elif ( len (append_obj) < 7 ) : # If attacked 1 time
			remain_count = 1
			total_stars = append_obj[4]
			append_obj.extend(('','',''))
		else :
			remain_count = 0
			total_stars = int(append_obj[4]) + int(append_obj[7])

		append_obj.append(remain_count)

		# Total stars

		append_obj.append(total_stars)
		
		# town hall
		append_obj.append(member['townhallLevel'])
		
		# Heroes level baba king
			

	
		### Best opponent attack info 
		if ( 'opponentAttacks' in member ) :
			#opponent_position = opponent_member[member['bestOpponentAttack']['attackerTag']]
			append_obj.append(opponent_member[member['bestOpponentAttack']['attackerTag']])
			append_obj.append(member['bestOpponentAttack']['stars'])
			append_obj.append(member['bestOpponentAttack']['destructionPercentage'])
			append_obj.append(member['opponentAttacks'])

		else : 
			append_obj.extend('','','','')

		append_xlsx('./',append_obj,title)

		append_obj = ['']

"""
