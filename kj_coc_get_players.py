#! /usr/bin/env python3


import requests, json
import datetime, time
import logging
import os
import openpyxl
from openpyxl.utils import get_column_letter



coc_id = '80dack@naver.com'
coc_passwd = 'dgenius11'
my_clan_tag='#982JCPJU'
my_key_name = "Created with coc.py Client" 
my_key = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6IjU2Yjc4NWQzLWRiNDYtNGQwZS04N2VhLWY3NGY1Y2RmN2MxMiIsImlhdCI6MTU4MzgwMzM1MCwic3ViIjoiZGV2ZWxvcGVyLzY5ZDVjYmRiLTg5NTQtMDc4Zi1hNzYwLTY2ZmY4MDFiMTc3MyIsInNjb3BlcyI6WyJjbGFzaCJdLCJsaW1pdHMiOlt7InRpZXIiOiJkZXZlbG9wZXIvc2lsdmVyIiwidHlwZSI6InRocm90dGxpbmcifSx7ImNpZHJzIjpbIjIyMC4yMzAuMTg0LjEzNSJdLCJ0eXBlIjoiY2xpZW50In1dfQ.N2ntG4IN3nuEOO5_rFtA0SRlo5nA2XDMdc3SxwyfcCfXBbB1T6yP29MHAsJri39zLylO14E7A_SvtHFcpDPYCA"

my_key2= "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6ImYzNjM2NDFiLWRkMTEtNDkwOS1hOGRmLTlhMzAyMmUyYmE2YSIsImlhdCI6MTU4NDQwNDg3OSwic3ViIjoiZGV2ZWxvcGVyLzY5ZDVjYmRiLTg5NTQtMDc4Zi1hNzYwLTY2ZmY4MDFiMTc3MyIsInNjb3BlcyI6WyJjbGFzaCJdLCJsaW1pdHMiOlt7InRpZXIiOiJkZXZlbG9wZXIvc2lsdmVyIiwidHlwZSI6InRocm90dGxpbmcifSx7ImNpZHJzIjpbIjE4Mi4yMjkuODAuMjAiXSwidHlwZSI6ImNsaWVudCJ9XX0.ehu1yYEbd7NkbuEJ6NGHF1AzRXcY9tTTEvbkr8xKdMlzTMqJ_tkq11LnBUsQwAefMaG-cPz0DvWJMRI3p0j7Rg"

API_PAGE_BASE = "https://developer.clashofclans.com/"
login_page = "https://developer.clashofclans.com/api/login"

login_data = {"email": coc_id, "password": coc_passwd}
headers = {"content-type": "application/json"}

response = requests.post(login_page, headers = headers, json = login_data)
#response_txt = json.loads(response.content.decode('utf-8'))

mylogger = logging.getLogger("kj_coc") #create kj_coc logging
mylogger.setLevel(logging.DEBUG) # more than DEBUG

#stream_handler = logging.StreamHandler()
#stream_handler.setFormatter(formatter)
formatter = logging.Formatter('%(asctime)s > %(module)s\t%(levelname)s\t%(message)s', '%Y-%m-%d %H:%M:%S')
#mylogger.setFormatter(formatter)
#mylogger.addHandler(stream_handler)

file_handler = logging.FileHandler('./kj_coc_log.txt')
file_handler.setFormatter(formatter)
file_handler.setLevel(logging.DEBUG) # more than DEBUG
mylogger.addHandler(file_handler)


def xlsx_idle_members(path, filename_xls) :

	if os.path.isfile(path+ filename_xls) :
		wb = openpyxl.load_workbook(path+filename_xls)
	else :
		wb = openpyxl.Workbook()

	sh = wb.create_sheet(str(datetime.datetime.now().date()), 0)
	sh['B2'] = 'Name'
	sh['C2'] = 'Townhall Level'
		
	#sh.merge_cells('C1:D1')
	sh['D2'] = 'Barbarian King Level'
	sh['E2'] = 'Archer Queen Level'
	sh['F2'] = 'Grand Warden Level'
	sh['G2'] = 'Royal Champion Level'
	sh.column_dimensions['B'].width = 17
	sh.column_dimensions['D'].width = 19
	sh.column_dimensions['E'].width = 19
	sh.column_dimensions['F'].width = 19
	sh.column_dimensions['G'].width = 19

	wb.save(path+filename_xls)

def append_xlsx(path, obj, filename_xls) :
	wb = openpyxl.load_workbook(path + filename_xls)
	sh = wb.active
	sh.append(obj)
	wb.save(path + filename_xls)

def get_player_with_file (player_tag) :
	api_players = 'https://api.clashofclans.com/v1/players/{playertag}'
	auth = { 'Authorization': 'Bearer {}'.format(my_key2) }
	request_url = (api_players.format(playertag=requests.utils.quote(player_tag)) )

	s = requests.Session()

	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %my_key2})

	now = datetime.datetime.now()
	current_time = now.strftime('%Y-%m-%d_%H_%M_%S')


	if ( res.status_code == 200) :
		res_txt = json.loads(res.content.decode('utf-8'))
		with open(current_time+'_members.json','w',encoding='utf-8') as members:
			json.dump(res_txt, members, indent="\t")
	else :
		# If not 200 OK
		pass
	
	return current_time+'_'+player_tag+'_members.json'


def get_player_no_file (player_tag) :
	api_players = 'https://api.clashofclans.com/v1/players/{playertag}'
	auth = { 'Authorization': 'Bearer {}'.format(my_key2) }
	request_url = (api_players.format(playertag=requests.utils.quote(player_tag)) )

	s = requests.Session()

	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %my_key2})

	if ( res.status_code == 200) :
		res_txt = json.loads(res.content.decode('utf-8'))
	else :
		# If not 200 OK
		pass
	
	return res_txt

def get_clan_members (clan_tag, key) :

	api_clan_members = 'https://api.clashofclans.com/v1/clans/{clan_tag}/members'

	auth = { 'Authorization': 'Bearer {}'.format(my_key) }
	request_url = (api_clan_members.format(clan_tag=requests.utils.quote(clan_tag)) )

	s = requests.Session()

	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %key})

	now = datetime.datetime.now()
	current_time = now.strftime('%Y-%m-%d_%H_%M_%S')


	if ( res.status_code == 200) :
		res_txt = json.loads(res.content.decode('utf-8'))
		with open(current_time+'_clan_members.json','w',encoding='utf-8') as members:
			json.dump(res_txt, members, indent="\t")

	else: 
		# If not 200 OK
		pass

	return current_time+'_clan_members.json'


def get_clan_war(clan_tag, key) :
	auth = { 'Authorization': 'Bearer {}'.format(key) }
	request_url = 'https://api.clashofclans.com/v1/clans/{clanTag}/currentwar'.format(clanTag = requests.utils.quote(clan_tag))
	s = requests.Session()

	res = s.get(request_url, headers={'Authorization': 'Bearer %s' %key})

	return res



def coc_login_kj ():

	login_page = "https://developer.clashofclans.com/api/login"

	login_data = {"email": coc_id, "password": coc_passwd}
	headers = {"content-type": "application/json"}

	response = requests.post(login_page, headers = headers, json = login_data)

	return response

if __name__ == "__main__":

	clan_members_json = get_clan_members(my_clan_tag, my_key2)
	#members_info_json = get_players(player_tag)

	add_list = ['']

	now = datetime.datetime.now()
	current_time = now.strftime('%Y-%m-%d_%H_%M')
	
	xlsx_file = current_time + '_players.xlsx'
	xlsx_idle_members('./', xlsx_file)			 


	#clan_members_json = '2020-03-30_10_34_46_clan_members.json'

	with open (clan_members_json, 'r', encoding='utf-8') as clan_members :
		json_data = json.load(clan_members)     

	for member in json_data['items'] :
		player_data = get_player_no_file(member['tag'])
		#file = get_player_no_file(member['tag'])

		player_name = player_data['name']
		player_townhall = player_data['townHallLevel']

		add_list.append(player_name)
		add_list.append(player_townhall)
		
		for hero in player_data['heroes'] :
		
			if (hero['name'] == 'Barbarian King') :
				player_king_level = hero['level']

			elif (hero['name'] == 'Archer Queen') :
				player_queen_level = hero['level']

			elif (hero['name'] == 'Grand Warden') :
				player_warden_level = hero['level']

			elif (hero['name'] == 'Royal Champion') :
				player_cham_level = hero['level']

		add_list.append(player_king_level)
		add_list.append(player_queen_level)
		add_list.append(player_warden_level)
		add_list.append(player_cham_level)

		append_xlsx('./', add_list, xlsx_file)	
		
		add_list = ['']
	

	
	


	
	


 



